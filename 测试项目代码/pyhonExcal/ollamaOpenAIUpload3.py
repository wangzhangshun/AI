#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import ollama
import json
import re
import hashlib
from openpyxl import load_workbook
from typing import List, Dict
from pathlib import Path
from diskcache import Cache


class RailwayPlanParser:
    def __init__(self, model_name: str = "deepseek-r1:7b"):
        """
        初始化施工计划解析器
        :param model_name: 使用的Ollama模型名称
        """
        self.client = ollama.Client()
        self.model = model_name
        self.cache = Cache("parser_cache")  # 缓存解析结果
        self.max_retries = 3

        # 字段映射表（可根据实际Excel调整）
        self.field_map = {
            "序号": "serialNumber",
            "日计划号": "planNumber",
            "月计划号": "monthPlanNumber",
            "线路": "lineName",
            "行别": "lineType",
            "施工项目": "homeworkTypes",
            "施工日期": "scheduleDate",
            "登记地点": "registerSite",
            "施工地点": "homeworkSite",
            "施工内容": "homeworkContent",
            "影响范围": "scopeInfluence",
            "施工单位及负责人": "homeworkUnit",
            "配合单位及负责人": "unitCoordination",
            "备注": "remarks"
        }

        # 输出JSON模板（用于指导模型）
        # self.output_template = {
        #     "serialNumber": 1,
        #     "planNumber": "R104150043",
        #     "lineType": "上行、下行",
        #     "homeworkTypes": ["电缆过道开挖", "轨道车收卸料"],
        #     "scheduleDate": "2025-04-15",
        #     "timePassage": "07:00:00",
        #     "downLaneTime": "10:40:00",
        #     "homeworkSite": "K297+000-K306+835",
        #     "homeworkContent": "1.电缆过道开挖 2.轨道车收卸料",
        #     "scopeInfluence": "封锁上行线297km-306km",
        #     "homeworkUnit": "河南东都电气工程有限公司",
        #     "homeworkTotalNumber": 25,
        #     "partyMembersNumber": 3,
        #     "homeworkOperations": ["张明杰", "梁占占"],
        #     "fieldProtectionOfficer": ["和佳伟", "李江涛"],
        #     "stationLiaisonOfficer": ["许艳芬", "杨瑞朋"],
        #     "unitCoordination": ["灵寿站工务", "行唐车站"],
        #     "keepEyeMan": "刘兰楷",
        #     "equipmentChange": "轨道车组4辆"
        # }

        self.output_template = {
            "serialNumber": 1,
            "planNumber": "R104150043",
            "lineType": "上行、下行",
            "homeworkTypes": "电缆过道开挖, 轨道车收卸料",
            "scheduleDate": "2025-04-15",
            "timePassage": "07:00:00",
            "downLaneTime": "10:40:00",
            "homeworkSite": "K297+000-K306+835",
            "homeworkContent": "1.电缆过道开挖 2.轨道车收卸料",
            "scopeInfluence": "封锁上行线297km-306km",
            "homeworkUnit": "河南东都电气工程有限公司",
            "homeworkTotalNumber": 25,
            "partyMembersNumber": 3,
            "homeworkOperations": "张明杰, 梁占占",
            "fieldProtectionOfficer": "和佳伟, 李江涛",
            "stationLiaisonOfficer": "许艳芬, 杨瑞朋",
            "unitCoordination": "灵寿站工务, 行唐车站",
            "keepEyeMan": "刘兰楷",
            "equipmentChange": "轨道车组4辆"
        }

    def parse(self, excel_path: str) -> List[Dict]:
        """
        解析Excel文件并返回标准化JSON
        :param excel_path: Excel文件路径
        :return: 解析结果列表
        """
        try:
            # 1. 读取并预处理Excel数据
            raw_data = self._read_excel(excel_path)
            if len(raw_data) < 2:
                return []

            # 2. 动态识别表头位置
            header_mapping = self._map_headers(raw_data[0])

            # 3. 处理每一行数据
            results = []
            for row in raw_data[1:]:
                if not any(row):  # 跳过空行
                    continue

                # 使用缓存避免重复处理
                row_hash = self._row_hash(row)
                if row_hash in self.cache:
                    results.append(self.cache[row_hash])
                    continue

                # 4. 构造大模型提示词
                prompt = self._build_prompt(row, header_mapping)

                # 5. 调用模型并解析结果
                response = self._query_model(prompt)
                parsed_data = self._parse_response(response)

                # 6. 后处理与缓存
                processed_data = self._post_process(parsed_data)
                self.cache[row_hash] = processed_data
                results.append(processed_data)

            return results

        except Exception as e:
            raise RuntimeError(f"解析过程中出错: {str(e)}")

    def _read_excel(self, path: str) -> List[List]:
        """读取Excel文件内容"""
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        return [list(row) for row in ws.iter_rows(values_only=True)]

    def _map_headers(self, header_row: List) -> Dict[str, int]:
        """动态映射表头到目标字段"""
        mapping = {}
        for idx, cell in enumerate(header_row):
            if not cell:
                continue
            cell_text = str(cell).lower().replace(" ", "")
            for field in self.field_map:
                if field.lower().replace(" ", "") in cell_text:
                    mapping[self.field_map[field]] = idx
        return mapping

    def _build_prompt(self, row_data: List, mapping: Dict[str, int]) -> str:
        """构造大模型提示词"""
        context = {}
        for field, idx in mapping.items():
            if idx < len(row_data):
                value = row_data[idx]
                context[field] = str(value) if value is not None else ""

        return f"""作为铁路施工计划专家，请将以下数据转换为标准JSON格式,planNumber或者lineType没有值，请忽略此行数据：

### 输入数据 ###
{json.dumps(context, ensure_ascii=False, indent=2)}

### 转换规则 ###
1. 时间处理：
   - "7:00" → "07:00:00"
   - "04月15日" → "2025-04-15"
2. 地点处理：
   - "297km000m" → "K297+000"
   - 合并连续里程标为范围表示
3. 人员提取：
   - 从"备注"中提取所有角色信息
4. 数值处理：
   - 自动统计作业总人数

### 输出要求 ###
必须严格按以下JSON格式输出，不要包含任何解释性文字：
```json
{json.dumps(self.output_template, ensure_ascii=False, indent=2)}
```"""

    def _query_model(self, prompt: str) -> str:
        """调用Ollama模型"""
        for attempt in range(self.max_retries):
            try:
                response = self.client.chat(
                    model=self.model,
                    messages=[{"role": "user", "content": prompt}],
                    options={
                        "temperature": max(0.1, 0.5 - attempt * 0.1),  # 逐步降低随机性
                        "num_ctx": 8192  # 增加上下文窗口
                    }
                )
                return response['message']['content']
            except Exception as e:
                if attempt == self.max_retries - 1:
                    raise RuntimeError(f"模型调用失败: {str(e)}")

    def _parse_response(self, response: str) -> Dict:
        """解析模型响应"""
        try:
            # 提取JSON部分
            json_match = re.search(r'```json\n?(.*?)\n?```', response, re.DOTALL)
            json_str = json_match.group(1) if json_match else response

            # 处理常见的JSON格式问题
            json_str = json_str.replace("'", '"')  # 单引号转双引号
            json_str = re.sub(r',\s*}', '}', json_str)  # 修复多余的逗号
            json_str = re.sub(r',\s*]', ']', json_str)

            return json.loads(json_str)
        except Exception as e:
            raise ValueError(f"响应解析失败: {str(e)}\n原始响应:\n{response}")

    def _post_process(self, data: Dict) -> Dict:
        """结果后处理"""
        # 确保必填字段存在
        required_fields = ["serialNumber", "planNumber", "homeworkSite"]
        for field in required_fields:
            if field not in data:
                data[field] = None

        # 类型转换
        if "serialNumber" in data:
            try:
                data["serialNumber"] = int(data["serialNumber"])
            except (ValueError, TypeError):
                data["serialNumber"] = 0

        if "homeworkTotalNumber" in data:
            try:
                data["homeworkTotalNumber"] = int(data["homeworkTotalNumber"] or 0)
            except (ValueError, TypeError):
                data["homeworkTotalNumber"] = 0

        # 自动从备注提取人员信息
        if "remarks" in data:
            data.update(self._extract_personnel(data["remarks"]))

        return data

    def _extract_personnel(self, remarks: str) -> Dict:
        """从备注提取人员信息"""
        result = {
            "homeworkOperations": [],
            "fieldProtectionOfficer": [],
            "stationLiaisonOfficer": [],
            "remoteGuard": []
        }

        if not remarks:
            return result

        # 提取各种角色
        patterns = {
            "homeworkOperations": r"(?:作业负责人|负责人)[:：]\s*(\S+)",
            "fieldProtectionOfficer": r"现场防护员[:：]\s*(\S+)",
            "stationLiaisonOfficer": r"驻站联络员[:：]\s*(\S+)",
            "remoteGuard": r"远端防护员[:：]\s*(\S+)"
        }

        for field, pattern in patterns.items():
            matches = re.findall(pattern, remarks)
            if matches:
                result[field] = matches

        # 提取盯控人
        if match := re.search(r"盯控人[:：]\s*(\S+)\s+(\d+)", remarks):
            result["keepEyeMan"] = f"{match.group(1)} {match.group(2)}"

        return result

    def _row_hash(self, row: List) -> str:
        """生成行数据哈希值"""
        row_str = "|".join(str(x) for x in row if x is not None)
        return hashlib.md5(row_str.encode()).hexdigest()

    def save_to_file(self, data: List[Dict], output_path: str):
        """保存结果到文件"""
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)


def main():
    import argparse

    parser = argparse.ArgumentParser(description='铁路施工计划解析工具')
    parser.add_argument('input', help='输入的Excel文件路径')
    parser.add_argument('-o', '--output', default='output.json', help='输出JSON文件路径')
    parser.add_argument('--model', default='deepseek-r1:7b', help='使用的Ollama模型名称')

    args = parser.parse_args()

    # 检查文件是否存在
    if not Path(args.input).exists():
        print(f"错误: 输入文件 {args.input} 不存在")
        return

    print(f"开始解析 {args.input} ...")

    try:
        # 初始化解析器
        plan_parser = RailwayPlanParser(model_name=args.model)

        # 执行解析
        result = plan_parser.parse(args.input)

        # 保存结果
        plan_parser.save_to_file(result, args.output)

        print(f"解析成功！结果已保存到 {args.output}")
        print(f"共解析 {len(result)} 条施工计划")

    except Exception as e:
        print(f"解析失败: {str(e)}")


if __name__ == "__main__":
    main()
    #python ollamaOpenAIUpload3.py .\ab.xlsx -o .\result.json