#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import requests
import json
import re
import hashlib
from openpyxl import load_workbook
from pathlib import Path
from typing import List, Dict, Optional
from urllib.parse import urljoin

class RemoteOllamaAdapter:
    """适配远程Ollama服务的调用"""

    def __init__(self, base_url: str = "http://localhost:11434", model: str = "deepseek-r1:7b"):
        self.base_url = base_url
        self.model = model
        self.session = requests.Session()
        self.timeout = 30

    def chat(self, prompt: str) -> str:
        """发送请求到远程Ollama服务"""
        url = urljoin(self.base_url, "/api/chat")
        payload = {
            "model": self.model,
            "messages": [{"role": "user", "content": prompt}],
            "options": {"temperature": 0.3},
            "stream": False  # 确保不返回流式响应
        }

        try:
            response = self.session.post(
                url,
                json=payload,
                timeout=self.timeout
            )
            response.raise_for_status()

            # 确保响应是完整的JSON
            response_data = response.json()
            if "message" in response_data and "content" in response_data["message"]:
                return response_data["message"]["content"]
            raise ValueError("Invalid response format from Ollama API")
        except requests.exceptions.RequestException as e:
            raise RuntimeError(f"Ollama API调用失败: {str(e)}")
        except json.JSONDecodeError as e:
            raise RuntimeError(f"Ollama API响应解析失败: {str(e)}")


class RailwayPlanParser:
    def __init__(self, ollama_url: str, model: str = "deepseek-r1:7b"):
        """初始化解析器"""
        self.ollama = RemoteOllamaAdapter(ollama_url, model)
        self.max_retries = 3

        # 字段映射配置
        self.field_map = {
            "序号": "serialNumber",
            "日计划号": "planNumber",
            "月计划号": "monthPlanNumber",
            "线路": "lineName",
            "行别": "lineType",
            "施工项目": "homeworkTypes",
            "施工日期": "scheduleDate",
            "登记地点": "registerSite",
            "施工地点": "homeworkSite",
            "施工内容": "homeworkContent",
            "影响范围": "scopeInfluence",
            "施工单位及负责人": "homeworkUnit",
            "配合单位及负责人": "unitCoordination",
            "备注": "remarks"
        }

        # 输出模板（指导模型生成）
        self.output_template = {
            "serialNumber": 1,
            "planNumber": "R104150043",
            "lineType": "上行、下行",
            "homeworkLevel": "Ⅲ",
            "homeworkTypes": "电缆过道开挖, 轨道车收卸料",
            "scheduleDate": "2025-04-15",
            "timePassage": "07:00:00",
            "downLaneTime": "10:40:00",
            "homeworkSite": "K297+000-K306+835",
            "homeworkContent": "1.电缆过道开挖 2.轨道车收卸料",
            "scopeInfluence": "封锁上行线297km-306km",
            "homeworkUnit": "河南东都电气工程有限公司",
            "homeworkTotalNumber": 25,
            "partyMembersNumber": 3,
            "homeworkOperations": "张明杰, 梁占占",
            "fieldProtectionOfficer": "和佳伟, 李江涛",
            "stationLiaisonOfficer": "许艳芬, 杨瑞朋",
            "unitCoordination": "灵寿站工务, 行唐车站",
            "keepEyeMan": "刘兰楷",
            "equipmentChange": "轨道车组4辆"
        }

    # def parse(self, excel_path: str) -> List[Dict]:
    #     """解析Excel文件"""
    #     try:
    #         # 1. 读取Excel数据
    #         raw_data = self._read_excel(excel_path)
    #         if len(raw_data) < 2:
    #             return []
    #
    #         # 2. 动态识别表头
    #         header_mapping = self._map_headers(raw_data[0])
    #
    #         # 3. 处理每一行数据
    #         results = []
    #         for row in raw_data[1:]:
    #             if not any(cell for cell in row if cell is not None):
    #                 continue
    #
    #             # 4. 构造提示词并调用模型
    #             prompt = self._build_prompt(row, header_mapping)
    #             print("调试 - 发送的提示词:", prompt)  # 调试用
    #
    #             response = self._query_ollama(prompt)
    #             print("调试 - 原始响应:", response)  # 调试用
    #
    #             parsed_data = self._parse_response(response)
    #             processed_data = self._post_process(parsed_data)
    #             results.append(processed_data)
    #
    #         return results
    #
    #     except Exception as e:
    #         raise RuntimeError(f"解析失败: {str(e)}")

    def parse(self, excel_path: str) -> List[Dict]:
        """解析Excel文件并返回去重后的数据"""
        try:
            raw_data = self._read_excel(excel_path)
            if len(raw_data) < 2:
                return []

            header_mapping = self._map_headers(raw_data[0])
            results = []
            seen_identifiers = set()  # 用于跟踪已处理的记录

            for row in raw_data[1:]:
                if not row or not any(cell for cell in row if cell and str(cell).strip()):
                    continue

                # 从原始数据中提取关键标识字段
                row_identifier = self._get_row_identifier(row, header_mapping)
                if row_identifier in seen_identifiers:
                    print(f"跳过重复行: {row_identifier}")
                    continue

                seen_identifiers.add(row_identifier)

                prompt = self._build_prompt(row, header_mapping)
                response = self._query_ollama(prompt)
                parsed_data = self._parse_response(response)
                processed_data = self._post_process(parsed_data)
                results.append(processed_data)

            return results

        except Exception as e:
            raise RuntimeError(f"解析失败: {str(e)}")

    def _get_row_identifier(self, row: List, mapping: Dict[str, int]) -> tuple:
        """从行数据中提取唯一标识"""
        serial = str(row[mapping.get("serialNumber", -1)]) if "serialNumber" in mapping else ""
        plan = str(row[mapping.get("planNumber", -1)]) if "planNumber" in mapping else ""
        return (serial, plan)  # 使用元组作为唯一键

    def _read_excel(self, path: str) -> List[List]:
        """读取Excel文件"""
        try:
            wb = load_workbook(path, data_only=True)
            ws = wb.active
            return [list(row) for row in ws.iter_rows(values_only=True)]
        except Exception as e:
            raise RuntimeError(f"读取Excel文件失败: {str(e)}")

    def _map_headers(self, header_row: List) -> Dict[str, int]:
        """动态映射表头位置"""
        mapping = {}
        for idx, cell in enumerate(header_row):
            if not cell:
                continue
            cell_text = str(cell).lower().replace(" ", "")
            for field in self.field_map:
                if field.lower().replace(" ", "") in cell_text:
                    mapping[self.field_map[field]] = idx
        return mapping

    def _build_prompt(self, row_data: List, mapping: Dict[str, int]) -> str:
        """构造大模型提示词"""
        context = {}
        for field, idx in mapping.items():
            if idx < len(row_data):
                value = row_data[idx]
                context[field] = str(value) if value is not None else ""

        return f"""作为铁路施工计划专家，请将以下数据转换为标准JSON格式：

### 输入数据 ###
{json.dumps(context, ensure_ascii=False, indent=2)}

### 转换规则 ###
1. 人员字段使用逗号分隔的字符串（不要用数组）
2. 时间格式："7:00" → "07:00:00"
3. 日期格式："04月15日" → "2025-04-15"
4. 里程格式："297km000m" → "K297+000"

### 输出要求 ###
必须严格按以下JSON格式输出（注意人员字段是字符串）：
```json
{json.dumps(self.output_template, ensure_ascii=False, indent=2)}
```"""

    def _query_ollama(self, prompt: str) -> str:
        """调用远程Ollama服务"""
        for attempt in range(self.max_retries):
            try:
                return self.ollama.chat(prompt)
            except Exception as e:
                if attempt == self.max_retries - 1:
                    raise
                continue

    def _parse_response(self, response: str) -> Dict:
        """解析模型响应"""
        try:
            # 尝试直接解析整个响应
            try:
                return json.loads(response)
            except json.JSONDecodeError:
                pass

            # 如果直接解析失败，尝试提取JSON部分
            json_match = re.search(r'```json\n?(.*?)\n?```', response, re.DOTALL)
            if json_match:
                json_str = json_match.group(1)
            else:
                # 尝试找到第一个{和最后一个}
                start = response.find('{')
                end = response.rfind('}')
                if start != -1 and end != -1:
                    json_str = response[start:end + 1]
                else:
                    raise ValueError("未找到有效的JSON数据")

            # 处理常见的JSON格式问题
            json_str = json_str.replace("'", '"')
            json_str = re.sub(r',\s*([}\]])', r'\1', json_str)
            json_str = re.sub(r'[\x00-\x1f\x7f-\x9f]', '', json_str)  # 移除控制字符

            data = json.loads(json_str)

            # 确保人员字段是字符串
            personnel_fields = [
                "homeworkTypes", "homeworkOperations",
                "fieldProtectionOfficer", "stationLiaisonOfficer",
                "unitCoordination"
            ]
            for field in personnel_fields:
                if field in data and isinstance(data[field], list):
                    data[field] = ", ".join(data[field])

            return data
        except Exception as e:
            raise ValueError(f"响应解析失败: {str(e)}\n原始响应:\n{response}")

    def _post_process(self, data: Dict) -> Dict:
        """结果后处理"""
        # 类型转换
        data["serialNumber"] = int(data.get("serialNumber", 0))
        data["homeworkTotalNumber"] = int(data.get("homeworkTotalNumber", 0))
        data["partyMembersNumber"] = int(data.get("partyMembersNumber", 0))

        # 从备注提取人员信息
        if "remarks" in data:
            personnel = self._extract_personnel(data["remarks"])
            for field, value in personnel.items():
                if field in data and data[field]:
                    data[field] += ", " + value
                else:
                    data[field] = value

        return data

    def _extract_personnel(self, remarks: str) -> Dict:
        """从备注提取人员信息"""
        result = {}
        if not remarks:
            return result

        # 提取规则
        patterns = {
            "homeworkOperations": r"(?:作业负责人|负责人)[:：]\s*(\S+)",
            "fieldProtectionOfficer": r"现场防护员[:：]\s*(\S+)",
            "stationLiaisonOfficer": r"驻站联络员[:：]\s*(\S+)",
            "unitCoordination": r"配合单位[:：]\s*([^\n]+)"
        }

        for field, pattern in patterns.items():
            matches = re.findall(pattern, remarks)
            if matches:
                result[field] = ", ".join(matches)

        # 提取盯控人
        if match := re.search(r"盯控人[:：]\s*(\S+)\s+(\d+)", remarks):
            result["keepEyeMan"] = f"{match.group(1)} {match.group(2)}"

        return result

    def save_to_file(self, data: List[Dict], output_path: str):
        """保存结果到文件"""
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)


def main():
    import argparse

    parser = argparse.ArgumentParser(description='铁路施工计划解析工具（远程Ollama版）')
    parser.add_argument('input', help='输入的Excel文件路径')
    parser.add_argument('--ollama', required=True, help='Ollama服务地址 (如 http://192.168.1.100:11434)')
    parser.add_argument('-o', '--output', default='output.json', help='输出JSON文件路径')
    parser.add_argument('--model', default='deepseek-r1:7b', help='使用的模型名称')

    args = parser.parse_args()

    if not Path(args.input).exists():
        print(f"错误: 输入文件 {args.input} 不存在")
        return

    print(f"开始解析 {args.input} (使用模型: {args.model})...")

    try:
        parser = RailwayPlanParser(args.ollama, args.model)
        result = parser.parse(args.input)
        parser.save_to_file(result, args.output)

        print(f"解析成功！结果已保存到 {args.output}")
        print(f"共解析 {len(result)} 条施工计划")

    except Exception as e:
        print(f"解析失败: {str(e)}")
        import traceback
        traceback.print_exc()  # 打印完整的错误堆栈


if __name__ == "__main__":
    main()