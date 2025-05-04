#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import requests
import json
import re
from openpyxl import load_workbook
from pathlib import Path
from typing import List, Dict
from urllib.parse import urljoin
from hashlib import md5


class RemoteOllamaAdapter:
    """适配远程Ollama服务的调用"""

    def __init__(self, base_url: str = "http://localhost:11434", model: str = "deepseek-r1:7b", api_key: str = None):
        self.base_url = base_url
        self.model = model
        self.session = requests.Session()
        self.timeout = 120
        self.api_key = api_key

    def chat(self, prompt: str) -> str:
        url = urljoin(self.base_url, "/api/chat")
        payload = {
            "model": self.model,
            "messages": [{"role": "user", "content": prompt}],
            "options": {"temperature": 0.3},
            "stream": False
        }

        headers = {}
        if self.api_key:
            headers["Authorization"] = f"Bearer {self.api_key}"

        try:
            response = self.session.post(url, json=payload, headers=headers, timeout=self.timeout)
            response.raise_for_status()
            return response.json()["message"]["content"]
        except Exception as e:
            raise RuntimeError(f"Ollama API调用失败: {str(e)}")


class RailwayPlanParser:
    def __init__(self, ollama_url: str, model: str = "deepseek-r1:7b", api_key: str = None):
        self.ollama = RemoteOllamaAdapter(ollama_url, model, api_key)
        self.max_retries = 3
        self.prompt_cache = {}
        self.batch_size = 10  # 每批处理的行数

        # 字段映射配置
        self.field_map = {
            "序号": "serialNumber",
            "日计划号": "planNumber",
            "月计划号": "monthPlanNumber",
            "线路": "lineName",
            "行别": "lineType",
            "施工项目": "homeworkTypes",
            "施工日期": "scheduleDate",
            "登记地点": "registerSite",
            "施工地点": "homeworkSite",
            "施工内容": "homeworkContent",
            "影响范围": "scopeInfluence",
            "施工单位及负责人": "homeworkUnit",
            "配合单位及负责人": "unitCoordination",
            "作业负责人": "homeworkOperations",
            "现场负责人": "siteLeader",
            "现场安全员": "siteSecurityOfficer",
            "现场防护员": "fieldProtectionOfficer",
            "驻站联络员": "stationLiaisonOfficer",
            "备注": "remarks"
        }

    def parse(self, excel_path: str) -> List[Dict]:
        """解析Excel文件"""
        try:
            # 1. 读取Excel数据
            raw_data = self._read_excel(excel_path)
            if len(raw_data) < 3:  # 表头+至少一行数据
                return []

            # 2. 动态识别表头
            header_mapping = self._map_headers(raw_data[1])  # 第2行是真正的表头

            # 3. 准备批量处理数据
            current_serial = 0
            all_rows = []

            # 收集所有有效行并处理序号
            for row in raw_data[2:]:
                if not self._is_valid_row(row):
                    continue

                # 处理序号列
                if row[0]:  # 如果序号列有值
                    current_serial = int(row[0])
                else:  # 如果序号列为空，使用上一个序号
                    row = list(row)
                    row[0] = current_serial

                all_rows.append(row)

            # 4. 分批处理
            results = []
            for i in range(0, len(all_rows), self.batch_size):
                batch_rows = all_rows[i:i + self.batch_size]
                try:
                    batch_result = self._process_batch(batch_rows, header_mapping)
                    results.extend(batch_result)
                except Exception as e:
                    print(f"批量处理失败: {str(e)}，回退到逐条处理")
                    for row in batch_rows:
                        try:
                            result = self._process_single_row(row, header_mapping)
                            results.append(result)
                        except Exception as e:
                            print(f"处理行数据失败(序号:{row[0]}): {str(e)}")
                            continue

            return results

        except Exception as e:
            raise RuntimeError(f"解析失败: {str(e)}")

    def _read_excel(self, path: str) -> List[List]:
        """读取Excel文件（修复样式问题）"""
        try:
            # 使用read_only模式并忽略样式
            wb = load_workbook(
                path,
                data_only=True,
                read_only=True,
                keep_links=False  # 忽略超链接
            )
            ws = wb.active

            # 获取最大行列数
            max_row = ws.max_row
            max_col = ws.max_column

            # 手动读取数据，避免样式问题
            data = []
            for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True):
                data.append(list(row))

            return data
        except Exception as e:
            # 如果仍然失败，尝试更基本的读取方式
            try:
                wb = load_workbook(path, data_only=True, read_only=True)
                ws = wb.active
                return [list(row) for row in ws.iter_rows(values_only=True)]
            except Exception as e:
                raise RuntimeError(f"读取Excel文件失败: {str(e)}")

    def _process_batch(self, rows: List[List], header_mapping: Dict[str, int]) -> List[Dict]:
        """处理一批数据"""
        batch_contexts = []
        for row in rows:
            context = {}
            for field, idx in header_mapping.items():
                if idx < len(row):
                    value = row[idx]
                    context[field] = str(value) if value is not None else ""
            batch_contexts.append(context)

        batch_prompt = self._build_batch_prompt(batch_contexts)
        batch_response = self._query_ollama(batch_prompt)
        parsed_results = self._parse_batch_response(batch_response)

        # 确保结果数量匹配
        if len(parsed_results) != len(rows):
            raise ValueError(f"结果数量不匹配: 预期 {len(rows)} 条，实际 {len(parsed_results)} 条")

        return [self._post_process(r) for r in parsed_results]

    def _process_single_row(self, row: List, header_mapping: Dict[str, int]) -> Dict:
        """处理单行数据"""
        prompt = self._build_prompt(row, header_mapping)
        response = self._query_ollama(prompt)
        parsed_data = self._parse_response(response)
        return self._post_process(parsed_data)

    def _build_batch_prompt(self, contexts: List[Dict]) -> str:
        """构建批量提示词"""
        template = {
            "serialNumber": 1,
            "planNumber": "R104150043",
            "lineType": "上行、下行",
            "homeworkLevel": "Ⅲ",
            "homeworkTypes": "电缆过道开挖,轨道车收卸料",
            "scheduleDate": "2025-04-15",
            "timePassage": "07:00:00",
            "downLaneTime": "10:40:00",
            "homeworkSite": "K297+000-K306+835",
            "homeworkContent": "1.电缆过道开挖 2.轨道车收卸料",
            "scopeInfluence": "封锁上行线297km-306km",
            "homeworkUnit": "河南东都电气工程有限公司",
            "homeworkTotalNumber": 25,
            "partyMembersNumber": 3,
            "homeworkOperations": "张明杰, 梁占占",
            "fieldProtectionOfficer": "和佳伟, 李江涛",
            "stationLiaisonOfficer": "许艳芬, 杨瑞朋",
            "unitCoordination": "灵寿站工务, 行唐车站",
            "keepEyeMan": "刘兰楷",
            "equipmentChange": "轨道车组4辆"
        }

        return f"""作为铁路施工计划专家，请将以下多组数据分别转换为标准JSON格式：

### 输入数据 ###
{json.dumps(contexts, ensure_ascii=False, indent=2)}

### 转换规则 ###
1. 人员字段使用逗号分隔的字符串
2. 时间格式："7:00" → "07:00:00"
3. 日期格式："04月15日" → "2025-04-15"
4. 里程格式："297km000m" → "K297+000"

### 输出要求 ###
必须严格按以下JSON数组格式输出，每个元素对应一组输入数据：
```json
{json.dumps([template], ensure_ascii=False, indent=2)}
```"""

    def _parse_batch_response(self, response: str) -> List[Dict]:
        """解析批量响应"""
        try:
            # 尝试提取JSON部分
            json_match = re.search(r'```json\n?(.*?)\n?```', response, re.DOTALL)
            json_str = json_match.group(1) if json_match else response

            # 处理常见JSON格式问题
            json_str = json_str.replace("'", '"')
            json_str = re.sub(r',\s*([}\]])', r'\1', json_str)
            return json.loads(json_str)
        except Exception as e:
            raise ValueError(f"批量响应解析失败: {str(e)}\n原始响应:\n{response}")

    def _is_valid_row(self, row: List) -> bool:
        """检查行是否有效"""
        return bool(row and any(cell and str(cell).strip() for cell in row))

    def _map_headers(self, header_row: List) -> Dict[str, int]:
        """动态映射表头位置"""
        mapping = {}
        for idx, cell in enumerate(header_row):
            if not cell:
                continue
            cell_text = str(cell).strip().lower().replace(" ", "")
            for field in self.field_map:
                if field.lower().replace(" ", "") in cell_text:
                    mapping[self.field_map[field]] = idx
        return mapping

    def _build_prompt(self, row_data: List, mapping: Dict[str, int]) -> str:
        """构造大模型提示词"""
        context = {}
        for field, idx in mapping.items():
            if idx < len(row_data):
                value = row_data[idx]
                context[field] = str(value) if value is not None else ""

        template = {
            "serialNumber": 1,
            "planNumber": "R104150043",
            "lineType": "上行、下行",
            "homeworkLevel": "Ⅲ",
            "homeworkTypes": "电缆过道开挖,轨道车收卸料",
            "scheduleDate": "2025-04-15",
            "timePassage": "07:00:00",
            "downLaneTime": "10:40:00",
            "homeworkSite": "K297+000-K306+835",
            "homeworkContent": "1.电缆过道开挖 2.轨道车收卸料",
            "scopeInfluence": "封锁上行线297km-306km",
            "homeworkUnit": "河南东都电气工程有限公司",
            "homeworkTotalNumber": 25,
            "partyMembersNumber": 3,
            "homeworkOperations": "张明杰, 梁占占",
            "fieldProtectionOfficer": "和佳伟, 李江涛",
            "stationLiaisonOfficer": "许艳芬, 杨瑞朋",
            "unitCoordination": "灵寿站工务, 行唐车站",
            "keepEyeMan": "刘兰楷",
            "equipmentChange": "轨道车组4辆"
        }

        return f"""作为铁路施工计划专家，请将以下数据转换为标准JSON格式：

### 输入数据 ###
{json.dumps(context, ensure_ascii=False, indent=2)}

### 转换规则 ###
1. 人员字段使用逗号分隔的字符串
2. 时间格式："7:00" → "07:00:00"
3. 日期格式："04月15日" → "2025-04-15"
4. 里程格式："297km000m" → "K297+000"

### 输出要求 ###
必须严格按以下JSON格式输出,所有字段值不能使用数组，如果为数组的请用逗号连接转成字符串：
```json
{json.dumps(template, ensure_ascii=False, indent=2)}
```"""

    def _query_ollama(self, prompt: str) -> str:
        """调用远程Ollama服务"""
        key = md5(prompt.encode("utf-8")).hexdigest()
        if key in self.prompt_cache:
            return self.prompt_cache[key]

        for attempt in range(self.max_retries):
            try:
                response = self.ollama.chat(prompt)
                self.prompt_cache[key] = response
                return response
            except Exception as e:
                if attempt == self.max_retries - 1:
                    raise

    def _parse_response(self, response: str) -> Dict:
        """解析模型响应"""
        try:
            # 尝试提取JSON部分
            json_match = re.search(r'```json\n?(.*?)\n?```', response, re.DOTALL)
            json_str = json_match.group(1) if json_match else response

            # 处理常见JSON格式问题
            json_str = json_str.replace("'", '"')
            json_str = re.sub(r',\s*([}\]])', r'\1', json_str)
            return json.loads(json_str)
        except Exception as e:
            raise ValueError(f"响应解析失败: {str(e)}\n原始响应:\n{response}")

    def _post_process(self, data: Dict) -> Dict:
        """结果后处理"""
        # 确保关键字段存在
        data.setdefault("homeworkTotalNumber", 0)
        data.setdefault("partyMembersNumber", 0)

        # 类型转换
        try:
            data["serialNumber"] = int(data.get("serialNumber", 0))
            data["homeworkTotalNumber"] = int(data.get("homeworkTotalNumber", 0))
            data["partyMembersNumber"] = int(data.get("partyMembersNumber", 0))
        except (ValueError, TypeError):
            pass

        return data

    def save_to_file(self, data: List[Dict], output_path: str):
        """保存结果到文件"""
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)


def main():
    import argparse

    parser = argparse.ArgumentParser(description='铁路施工计划解析工具')
    parser.add_argument('input', help='输入的Excel文件路径')
    parser.add_argument('--ollama', required=True, help='Ollama服务地址')
    parser.add_argument('-o', '--output', default='output.json', help='输出JSON文件路径')
    parser.add_argument('--model', default='deepseek-r1:7b', help='使用的模型名称')

    args = parser.parse_args()

    if not Path(args.input).exists():
        print(f"错误: 输入文件 {args.input} 不存在")
        return

    print(f"开始解析 {args.input}...")

    try:
        parser = RailwayPlanParser(args.ollama, args.model)
        result = parser.parse(args.input)
        parser.save_to_file(result, args.output)
        print(f"解析成功！共解析 {len(result)} 条记录")
        print(f"结果已保存到 {args.output}")

    except Exception as e:
        print(f"解析失败: {str(e)}")


if __name__ == "__main__":
    main()