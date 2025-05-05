#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import requests
import json
import re
import time
import logging
from openpyxl import load_workbook
from pathlib import Path
from typing import List, Dict, Optional
from urllib.parse import urljoin
from hashlib import md5
from tenacity import retry, stop_after_attempt, wait_fixed, retry_if_exception_type

# 配置日志
logger = logging.getLogger(__name__)


class RemoteOllamaAdapter:
    """适配远程Ollama服务的调用"""

    def __init__(self, base_url: str = "http://localhost:11434",
                 model: str = "deepseek-r1:7b",
                 api_key: str = None):
        self.base_url = base_url
        self.model = model
        self.session = requests.Session()
        self.timeout = 600  # 10分钟超时
        self.api_key = api_key

        # 配置会话
        self.session.headers.update({
            "Content-Type": "application/json",
            "Accept": "application/json"
        })
        if self.api_key:
            self.session.headers.update({"Authorization": f"Bearer {self.api_key}"})

    @retry(
        stop=stop_after_attempt(3),
        wait=wait_fixed(5),
        retry=retry_if_exception_type(requests.exceptions.RequestException)
    )
    def chat(self, prompt: str) -> str:
        url = urljoin(self.base_url, "/api/chat")
        payload = {
            "model": self.model,
            "messages": [{"role": "user", "content": prompt}],
            "options": {"temperature": 0.3},
            "stream": False
        }

        try:
            logger.debug("发送请求到Ollama API，模型: %s", self.model)
            start_time = time.time()

            response = self.session.post(
                url,
                json=payload,
                timeout=self.timeout
            )
            response.raise_for_status()

            process_time = time.time() - start_time
            logger.debug("Ollama API响应时间: %.2f秒", process_time)

            return response.json()["message"]["content"]

        except requests.exceptions.Timeout:
            logger.error("Ollama API请求超时")
            raise RuntimeError("Ollama API请求超时，请稍后重试")
        except Exception as e:
            logger.error("Ollama API调用失败: %s", str(e))
            raise RuntimeError(f"Ollama API调用失败: {str(e)}")


class RailwayPlanParser:
    def __init__(self, ollama_url: str, model: str = "deepseek-r1:7b", api_key: str = None):
        self.ollama = RemoteOllamaAdapter(ollama_url, model, api_key)
        self.max_retries = 3
        self.prompt_cache = {}
        self.batch_size = 5  # 初始批量大小
        self.max_batch_size = 20  # 最大批量大小
        self.min_batch_size = 1  # 最小批量大小
        self.cache_size_limit = 100  # 缓存大小限制

        # 字段映射配置
        self.field_map = {
            "序号": "serialNumber",
            "日计划号": "planNumber",
            "月计划号": "monthPlanNumber",
            "线路": "lineName",
            "行别": "lineType",
            "施工项目": "homeworkTypes",
            "施工日期": "scheduleDate",
            "登记地点": "registerSite",
            "施工地点": "homeworkSite",
            "施工内容": "homeworkContent",
            "影响范围": "scopeInfluence",
            "施工单位及负责人": "homeworkUnit",
            "配合单位及负责人": "unitCoordination",
            "作业负责人": "homeworkOperations",
            "现场负责人": "siteLeader",
            "现场安全员": "siteSecurityOfficer",
            "现场防护员": "fieldProtectionOfficer",
            "驻站联络员": "stationLiaisonOfficer",
            "备注": "remarks"
        }

        self.logger = logging.getLogger(f"{__name__}.RailwayPlanParser")

    def parse(self, excel_path: str) -> List[Dict]:
        """解析Excel文件"""
        self.logger.info("开始解析Excel文件: %s", excel_path)
        start_time = time.time()

        try:
            # 1. 读取Excel数据
            raw_data = self._read_excel(excel_path)
            if len(raw_data) < 3:  # 表头+至少一行数据
                self.logger.warning("Excel数据行数不足: %d", len(raw_data))
                return []

            # 2. 动态识别表头
            header_mapping = self._map_headers(raw_data[1])  # 第2行是真正的表头
            self.logger.debug("表头映射: %s", header_mapping)

            # 3. 准备批量处理数据
            current_serial = 0
            all_rows = []

            # 收集所有有效行并处理序号
            for row in raw_data[2:]:
                if not self._is_valid_row(row):
                    continue

                # 处理序号列
                if row[0]:  # 如果序号列有值
                    try:
                        current_serial = int(row[0])
                    except (ValueError, TypeError):
                        pass
                else:  # 如果序号列为空，使用上一个序号
                    row = list(row)
                    row[0] = current_serial

                all_rows.append(row)

            self.logger.info("共发现 %d 行有效数据", len(all_rows))

            # 4. 分批处理
            results = []
            for i in range(0, len(all_rows), self.batch_size):
                batch_rows = all_rows[i:i + self.batch_size]
                try:
                    batch_result = self._process_batch(batch_rows, header_mapping)
                    results.extend(batch_result)

                    # 动态调整批量大小
                    if len(batch_rows) == self.batch_size and self.batch_size < self.max_batch_size:
                        self.batch_size = min(self.batch_size + 2, self.max_batch_size)
                        self.logger.debug("增加批量大小到 %d", self.batch_size)

                except Exception as e:
                    self.logger.warning("批量处理失败: %s，回退到逐条处理", str(e))
                    self.batch_size = max(self.batch_size - 2, self.min_batch_size)
                    self.logger.debug("减小批量大小到 %d", self.batch_size)

                    for row in batch_rows:
                        try:
                            result = self._process_single_row(row, header_mapping)
                            results.append(result)
                        except Exception as e:
                            self.logger.error("处理行数据失败(序号:%s): %s", row[0], str(e))
                            continue

            process_time = time.time() - start_time
            self.logger.info("解析完成，共解析 %d 条记录，耗时 %.2f 秒", len(results), process_time)
            return results

        except Exception as e:
            self.logger.error("解析失败: %s", str(e))
            raise RuntimeError(f"解析失败: {str(e)}")

    def _read_excel(self, path: str) -> List[List]:
        """读取Excel文件"""
        self.logger.debug("开始读取Excel文件: %s", path)
        data = []
        wb = None
        try:
            # 不使用上下文管理器，手动管理资源
            wb = load_workbook(
                path,
                data_only=True,
                read_only=True,
                keep_links=False
            )
            ws = wb.active
            max_row = ws.max_row
            max_col = ws.max_column

            for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True):
                data.append(list(row))

            self.logger.debug("成功读取 %d 行数据", len(data))
            return data

        except Exception as e:
            self.logger.error("读取Excel文件失败: %s", str(e))
            raise RuntimeError(f"读取Excel文件失败: {str(e)}")
        finally:
            # 手动关闭工作簿
            if wb:
                wb.close()

    def _process_batch(self, rows: List[List], header_mapping: Dict[str, int]) -> List[Dict]:
        """处理一批数据"""
        self.logger.debug("开始批量处理 %d 行数据", len(rows))
        start_time = time.time()
        remaining_attempts = self.max_retries

        while remaining_attempts > 0:
            try:
                batch_contexts = []
                for row in rows:
                    context = {}
                    for field, idx in header_mapping.items():
                        if idx < len(row):
                            value = row[idx]
                            context[field] = str(value) if value is not None else ""
                    batch_contexts.append(context)

                batch_prompt = self._build_batch_prompt(batch_contexts)
                batch_response = self._query_ollama(batch_prompt)
                parsed_results = self._parse_batch_response(batch_response)

                if len(parsed_results) == len(rows):
                    process_time = time.time() - start_time
                    self.logger.debug("批量处理成功，耗时 %.2f 秒", process_time)
                    return [self._post_process(r) for r in parsed_results]

                # 如果结果数量不匹配，尝试减小批量大小
                if len(rows) > 1:
                    self.logger.warning("结果数量不匹配 (预期: %d, 实际: %d)，尝试减小批量大小",
                                        len(rows), len(parsed_results))
                    mid = len(rows) // 2
                    first_half = self._process_batch(rows[:mid], header_mapping)
                    second_half = self._process_batch(rows[mid:], header_mapping)
                    return first_half + second_half

                remaining_attempts -= 1
                if remaining_attempts > 0:
                    time.sleep(5)  # 重试前等待

            except Exception as e:
                remaining_attempts -= 1
                if remaining_attempts > 0:
                    self.logger.warning("批量处理失败，剩余重试次数: %d, 错误: %s",
                                        remaining_attempts, str(e))
                    time.sleep(5)
                else:
                    raise RuntimeError(f"批量处理失败: {str(e)}")

        # 如果所有重试都失败，回退到逐条处理
        self.logger.warning("批量处理失败，回退到逐条处理")
        results = []
        for row in rows:
            try:
                result = self._process_single_row(row, header_mapping)
                results.append(result)
            except Exception as e:
                self.logger.error("处理行数据失败(序号:%s): %s", row[0], str(e))
                continue
        return results

    def _process_single_row(self, row: List, header_mapping: Dict[str, int]) -> Dict:
        """处理单行数据"""
        prompt = self._build_prompt(row, header_mapping)
        response = self._query_ollama(prompt)
        parsed_data = self._parse_response(response)
        return self._post_process(parsed_data)

    def _build_batch_prompt(self, contexts: List[Dict]) -> str:
        """构建批量提示词"""
        template = {
            "serialNumber": 1,
            "planNumber": "R104150043",
            "lineType": "上行、下行",
            "homeworkLevel": "Ⅲ",
            "homeworkTypes": "电缆过道开挖,轨道车收卸料",
            "scheduleDate": "2025-04-15",
            "timePassage": "07:00:00",
            "downLaneTime": "10:40:00",
            "homeworkSite": "K297+000-K306+835",
            "homeworkContent": "1.电缆过道开挖 2.轨道车收卸料",
            "scopeInfluence": "封锁上行线297km-306km",
            "homeworkUnit": "河南东都电气工程有限公司",
            "homeworkTotalNumber": 25,
            "partyMembersNumber": 3,
            "homeworkOperations": "张明杰, 梁占占",
            "fieldProtectionOfficer": "和佳伟, 李江涛",
            "stationLiaisonOfficer": "许艳芬, 杨瑞朋",
            "unitCoordination": "灵寿站工务, 行唐车站",
            "keepEyeMan": "刘兰楷",
            "equipmentChange": "轨道车组4辆"
        }

        return f"""作为铁路施工计划专家，请将以下多组数据分别转换为标准JSON格式：

### 输入数据 ###
{json.dumps(contexts, ensure_ascii=False, indent=2)}

### 转换规则 ###
1. 人员字段使用逗号分隔的字符串
2. 时间格式："7:00" → "07:00:00"
3. 日期格式："04月15日" → "2025-04-15"
4. 里程格式："297km000m" → "K297+000"

### 输出要求 ###
必须严格按以下JSON数组格式输出，每个元素对应一组输入数据：
```json
{json.dumps([template], ensure_ascii=False, indent=2)}
```"""

    def _parse_batch_response(self, response: str) -> List[Dict]:
        """解析批量响应"""
        try:
            # 尝试提取JSON部分
            json_match = re.search(r'(?s)\[.*\]', response)
            json_str = json_match.group(0) if json_match else response

            # 处理常见JSON格式问题
            json_str = json_str.replace("'", '"')
            json_str = re.sub(r',\s*([}\]])', r'\1', json_str)
            json_str = re.sub(r'([{,])\s*([^"\s{}][^:\s{}]*)\s*:', r'\1"\2":', json_str)

            return json.loads(json_str)
        except Exception as e:
            self.logger.error("批量响应解析失败: %s", str(e))
            raise ValueError(f"批量响应解析失败: {str(e)}\n原始响应:\n{response[:500]}...")

    def _is_valid_row(self, row: List) -> bool:
        """检查行是否有效"""
        return bool(row and any(cell and str(cell).strip() for cell in row))

    def _map_headers(self, header_row: List) -> Dict[str, int]:
        """动态映射表头位置"""
        mapping = {}
        for idx, cell in enumerate(header_row):
            if not cell:
                continue
            cell_text = str(cell).strip().lower().replace(" ", "")
            for field in self.field_map:
                if field.lower().replace(" ", "") in cell_text:
                    mapping[self.field_map[field]] = idx
        return mapping

    def _build_prompt(self, row_data: List, mapping: Dict[str, int]) -> str:
        """构造大模型提示词"""
        context = {}
        for field, idx in mapping.items():
            if idx < len(row_data):
                value = row_data[idx]
                context[field] = str(value) if value is not None else ""

        template = {
            "serialNumber": 1,
            "planNumber": "R104150043",
            "lineType": "上行、下行",
            "homeworkLevel": "Ⅲ",
            "homeworkTypes": "电缆过道开挖,轨道车收卸料",
            "scheduleDate": "2025-04-15",
            "timePassage": "07:00:00",
            "downLaneTime": "10:40:00",
            "homeworkSite": "K297+000-K306+835",
            "homeworkContent": "1.电缆过道开挖 2.轨道车收卸料",
            "scopeInfluence": "封锁上行线297km-306km",
            "homeworkUnit": "河南东都电气工程有限公司",
            "homeworkTotalNumber": 25,
            "partyMembersNumber": 3,
            "homeworkOperations": "张明杰, 梁占占",
            "fieldProtectionOfficer": "和佳伟, 李江涛",
            "stationLiaisonOfficer": "许艳芬, 杨瑞朋",
            "unitCoordination": "灵寿站工务, 行唐车站",
            "keepEyeMan": "刘兰楷",
            "equipmentChange": "轨道车组4辆"
        }

        return f"""作为铁路施工计划专家，请将以下数据转换为标准JSON格式：

### 输入数据 ###
{json.dumps(context, ensure_ascii=False, indent=2)}

### 转换规则 ###
1. 人员字段使用逗号分隔的字符串
2. 时间格式："7:00" → "07:00:00"
3. 日期格式："04月15日" → "2025-04-15"
4. 里程格式："297km000m" → "K297+000"

### 输出要求 ###
必须严格按以下JSON格式输出,所有字段值不能使用数组，如果为数组的请用逗号连接转成字符串：
```json
{json.dumps(template, ensure_ascii=False, indent=2)}
```"""

    def _query_ollama(self, prompt: str) -> str:
        """调用远程Ollama服务"""
        key = md5(prompt.encode("utf-8")).hexdigest()
        if key in self.prompt_cache:
            self.logger.debug("使用缓存响应")
            return self.prompt_cache[key]

        for attempt in range(self.max_retries):
            try:
                response = self.ollama.chat(prompt)

                # 管理缓存大小
                if len(self.prompt_cache) >= self.cache_size_limit:
                    self.prompt_cache.pop(next(iter(self.prompt_cache)))

                self.prompt_cache[key] = response
                return response
            except Exception as e:
                if attempt == self.max_retries - 1:
                    self.logger.error("Ollama查询失败: %s", str(e))
                    raise
                time.sleep(3)

    def _parse_response(self, response: str) -> Dict:
        """解析模型响应"""
        try:
            # 尝试直接解析
            try:
                return json.loads(response.strip())
            except json.JSONDecodeError:
                pass

            # 尝试提取JSON部分
            json_match = re.search(r'(?s)\{.*\}', response)
            json_str = json_match.group(0) if json_match else response

            # 处理常见JSON格式问题
            json_str = json_str.replace("'", '"')
            json_str = re.sub(r',\s*([}\]])', r'\1', json_str)
            json_str = re.sub(r'([{,])\s*([^"\s{}][^:\s{}]*)\s*:', r'\1"\2":', json_str)

            return json.loads(json_str)
        except Exception as e:
            self.logger.error("响应解析失败: %s", str(e))
            raise ValueError(f"响应解析失败: {str(e)}\n原始响应:\n{response[:500]}...")

    def _post_process(self, data: Dict) -> Dict:
        """结果后处理"""
        # 确保关键字段存在
        data.setdefault("homeworkTotalNumber", 0)
        data.setdefault("partyMembersNumber", 0)

        # 类型转换
        try:
            data["serialNumber"] = int(data.get("serialNumber", 0))
            data["homeworkTotalNumber"] = int(data.get("homeworkTotalNumber", 0))
            data["partyMembersNumber"] = int(data.get("partyMembersNumber", 0))
        except (ValueError, TypeError):
            pass

        return data

    def save_to_file(self, data: List[Dict], output_path: str):
        """保存结果到文件"""
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)


def main():
    import argparse

    parser = argparse.ArgumentParser(description='铁路施工计划解析工具')
    parser.add_argument('input', help='输入的Excel文件路径')
    parser.add_argument('--ollama', required=True, help='Ollama服务地址')
    parser.add_argument('-o', '--output', default='output.json', help='输出JSON文件路径')
    parser.add_argument('--model', default='deepseek-r1:7b', help='使用的模型名称')

    args = parser.parse_args()

    if not Path(args.input).exists():
        print(f"错误: 输入文件 {args.input} 不存在")
        return

    print(f"开始解析 {args.input}...")

    try:
        parser = RailwayPlanParser(args.ollama, args.model)
        result = parser.parse(args.input)
        parser.save_to_file(result, args.output)
        print(f"解析成功！共解析 {len(result)} 条记录")
        print(f"结果已保存到 {args.output}")

    except Exception as e:
        print(f"解析失败: {str(e)}")


if __name__ == "__main__":
    main()